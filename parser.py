
import openpyxl
import json

# odpremo excel file
workbook = openpyxl.load_workbook('proba.xlsx')
worksheet = workbook.active

stVrstic = worksheet.max_row
stStolp = worksheet.max_column

results = {}
povprecjeResults = {}

# Gremo cez vse vrstice in stolpce in zapisujemo vrednosti v results
for i in range(2, stVrstic + 1):
    trenVrstica = worksheet.cell(row=i, column=1).value
    results[trenVrstica] = {}
    for j in range(2, stStolp + 1):
        trenStolp = worksheet.cell(row=1, column=j).value
        trenVrednostCelice = worksheet.cell(row=i, column=j).value
        results[trenVrstica][trenStolp] = trenVrednostCelice



# vpisemo vse iz results v json file
with open('results.json', 'w') as f:
    json.dump(results, f, indent=4)


# npr da izracunamo povprecje vsake vrstice in vpisemo v povprecje.json
for i in range(2, stVrstic + 1):
    trenVrsticaList = []
    for j in range(2, stStolp + 1):
        trenVrednostCelice = worksheet.cell(row=i, column=j).value
        if trenVrednostCelice:
            trenVrsticaList.append(trenVrednostCelice)
    if trenVrsticaList:
        avg = sum(trenVrsticaList) / len(trenVrsticaList)
        povprecjeResults[f"vrstica:{i-1}"] = avg

 
with open('povprecje.json', 'w') as f:
    json.dump(povprecjeResults, f, indent=4)